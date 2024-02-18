from deepdiff import DeepDiff
from dateutil.relativedelta import relativedelta
from urllib.parse import urlparse
import datetime
import openpyxl
import requests


class PytestFunctions:

    # 读取xlsx并转为LIST_DICT
    def readXlsx(self, file_path, sheet_name='Sheet1', head=True) -> list:
        """文件路径，sheet页，是否有列名"""
        myxls = openpyxl.load_workbook(file_path)
        activeSheet = myxls[sheet_name]
        if head:
            keys, xlsxData = list(), list()
            for row in range(1, activeSheet.max_row + 1):
                d = dict()
                for n, column in enumerate(range(1, activeSheet.max_column + 1)):
                    data = activeSheet.cell(row=row, column=column).value
                    if len(keys) < activeSheet.max_column:
                        keys.append(data)
                    else:
                        d[keys[n]] = data
                if d:
                    xlsxData.append(d)
        else:
            xlsxData = list()
            for row in range(1, activeSheet.max_row + 1):
                xlsxData.append(
                    [activeSheet.cell(row=row, column=column).value for column in range(1, activeSheet.max_column + 1)])
        return xlsxData

    # 日期处理
    def timeShift(self, strftime=None, **kwargs):
        '''
        支持：years, months, days, weeks, hours, minutes, seconds, microseconds
        例子1： timeShift("%Y-%m-%d %H:%M:%S", days=1) # 明天的当前时间
        例子2： timeShift("%Y-%m-%d", month=1, day=1)  # 今年的1月1日
        '''
        new_date = datetime.datetime.now() + relativedelta(**kwargs)
        if strftime:
            new_date = new_date.strftime(strftime)
        return new_date

    # 数据比对
    def dataDiff(self, data1, data2, **kwargs):
        diff = DeepDiff(data1, data2, **kwargs)
        if diff:
            raise AssertionError(diff.pretty())

    # url元组
    def getUrl(self, url):
        return urlparse(url)

    # # 字典cookies转cookiejar格式
    # def toCookiejar(self, data: dict):
    #     return requests.utils.cookiejar_from_dict(data)