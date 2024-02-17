import os
from pathlib import Path
from typing import Dict

import numpy as np
import datetime
import shutil
import openpyxl
import json
import ruamel.yaml

from steam_sdk.builders.BuilderLEDET import BuilderLEDET
from steam_sdk.data.DataLEDET import LEDETInputs, LEDETOptions, LEDETPlots, LEDETVariables
from steam_sdk.data.TemplateLEDET import get_template_LEDET_inputs_sheet, get_template_LEDET_options_sheet, get_template_LEDET_plots_sheet, get_template_LEDET_variables_sheet
from steam_sdk.parsers import ParserMap2d
from steam_sdk.parsers.ParserExcel import read_row, write2Excel
from steam_sdk.parsers.ParserYAML import dict_to_yaml
from steam_sdk.utils.NumpyEncoder import NumpyEncoder
from steam_sdk.utils.rgetattr import rgetattr
from steam_sdk.utils.sgetattr import rsetattr
from steam_sdk.utils.compare_two_parameters import compare_two_parameters


class ParserLEDET:
    """
        Class with methods to read/write LEDET information from/to other programs
    """

    def __init__(self, builder_ledet: BuilderLEDET):
        """
            Initialization using a BuilderLEDET object containing LEDET parameter structure
        """

        self.builder_ledet: BuilderLEDET = builder_ledet


    def readFromExcel(self, file_name: str, verbose: bool = True):
        '''
            *** Function that reads an Excel file defining a LEDET input file and imports it in a BuilderLEDET object ***

            :param file_name: Name of the file to read
            :type file_name: str
            :param verbose: Flag indicating whether additional information should be displayed
            :type verbose: str

            :return: None
        '''

        # Unpack variables
        builder_ledet = self.builder_ledet

        ##File must be whole eos string
        workbookVariables = openpyxl.load_workbook(file_name, data_only=True)

        #Inputs
        worksheetInputs = workbookVariables['Inputs']
        lastAttribute = worksheetInputs.cell(1, 2).value
        for i in range(1, worksheetInputs.max_row+1):
            # builder_ledet.variablesInputs[str(worksheetInputs.cell(i, 2).value)] = str(worksheetInputs.cell(i, 1).value)
            attribute = worksheetInputs.cell(i, 2).value
            try:
                if (attribute == None):
                    if worksheetInputs.cell(i, 3).value is not None:
                        values = read_row(worksheetInputs, i)
                        values = np.array([k for k in values if(str(k))])
                        current = builder_ledet.getAttribute(builder_ledet.Inputs, lastAttribute)
                        current = np.vstack((current, values))
                        builder_ledet.setAttribute(builder_ledet.Inputs, lastAttribute, current)
                    else:
                        continue
                elif (type(builder_ledet.getAttribute(builder_ledet.Inputs, attribute)) == np.ndarray):
                    lastAttribute = attribute
                    values = read_row(worksheetInputs, i)
                    values = np.array([k for k in values if(str(k))])
                    builder_ledet.setAttribute(builder_ledet.Inputs, attribute, values)
                else:
                    value = worksheetInputs.cell(i, 3).value
                    builder_ledet.setAttribute(builder_ledet.Inputs, attribute, value)
            except TypeError as e:
                if attribute in builder_ledet.Inputs.__annotations__: raise e
                if attribute=='None' or attribute==None: continue
                if verbose: print("Error with attribute: {}, continuing.".format(attribute))
        #Options
        worksheetOptions = workbookVariables['Options']
        for i in range(1, worksheetOptions.max_row+1):
            # builder_ledet.variablesOptions[str(worksheetOptions.cell(i, 2).value)] = str(worksheetOptions.cell(i, 1).value)
            attribute = worksheetOptions.cell(i, 2).value
            try:
                if (type(builder_ledet.getAttribute(builder_ledet.Options, attribute)) == np.ndarray):
                    values = read_row(worksheetOptions, i)
                    values = np.array([k for k in values if(str(k))])
                    builder_ledet.setAttribute(builder_ledet.Options, attribute, values)
                else:
                    value = worksheetOptions.cell(i, 3).value
                    builder_ledet.setAttribute(builder_ledet.Options, attribute, value)
            except TypeError as e:
                if attribute in builder_ledet.Options.__annotations__: raise e
                if attribute == 'None' or attribute == None: continue
                if verbose: print("Error with attribute: {}, continuing.".format(attribute))
        #Plots
        worksheetPlots = workbookVariables['Plots']
        for i in range(1, worksheetPlots.max_row+1):
            # builder_ledet.variablesPlots[str(worksheetPlots.cell(i, 2).value)] = str(worksheetPlots.cell(i, 1).value)
            attribute = worksheetPlots.cell(i, 2).value
            try:
                if (type(builder_ledet.getAttribute(builder_ledet.Plots, attribute)) == np.ndarray):
                    values = read_row(worksheetPlots, i, St=True)[2:]
                    values = np.array([k for k in values if(str(k))])
                    builder_ledet.setAttribute(builder_ledet.Plots, attribute, values)
                else:
                    try:
                        value = worksheetPlots.cell(i, 3).value
                    except:
                        value = ''
                    builder_ledet.setAttribute(builder_ledet.Plots, attribute, value)
            except TypeError as e:
                if attribute == 'None' or attribute == None: continue
                if verbose: print("Error with attribute: {}, continuing.".format(attribute))
        # Variables
        try:
            worksheetVariables = workbookVariables['Variables']
            for i in range(1, worksheetVariables.max_row+1):
                # builder_ledet.variablesVariables[str(worksheetVariables.cell(i, 2).value)] = str(worksheetVariables.cell(i, 1).value)
                attribute = worksheetVariables.cell(i, 2).value
                try:
                    if (type(builder_ledet.getAttribute(builder_ledet.Variables, attribute)) == np.ndarray):
                        if attribute != 'typeVariableToSaveTxt':  values = read_row(worksheetVariables, i, St = True)[2:]
                        else:  values = read_row(worksheetVariables, i)
                        values = np.array([k for k in values if(str(k))])
                        builder_ledet.setAttribute(builder_ledet.Variables, attribute, values)
                    else:
                        value = worksheetVariables.cell(i, 3).value
                        builder_ledet.setAttribute(builder_ledet.Variables, attribute, value)
                except TypeError as e:
                    if attribute in builder_ledet.Variables.__annotations__: raise e
                    if attribute == 'None' or attribute == None: continue
                    if verbose: print("Error with attribute: {}, continuing.".format(attribute))
        except:
            pass
            print("Error while reading Variables. Please check!")

        return builder_ledet


    def writeLedet2Excel(self, full_path_file_name: str, verbose: bool = False, SkipConsistencyCheck: bool = True):
        '''
        ** Writes a LEDET input file **

        :param full_path_file_name:
        :param verbose:
        :param SkipConsistencyCheck:
        :return:
        '''

        self._expand_scalar_to_array(verbose)

        # Import templates for LEDET input file sheets
        template_LEDET_inputs_sheet    = get_template_LEDET_inputs_sheet()
        template_LEDET_options_sheet   = get_template_LEDET_options_sheet()
        template_LEDET_plots_sheet     = get_template_LEDET_plots_sheet()
        template_LEDET_variables_sheet = get_template_LEDET_variables_sheet()

        # Define optional variables, which will be written only if present in the dataclass
        optional_variables_input_sheet  = _getOptionalVariables_input()
        optional_variables_output_sheet = _getOptionalVariables_output()

        ### Inputs sheet
        name_sheet_inputs = 'Inputs'  # This defines the name of the sheet and also of the variable group
        LEDET_inputs_sheet = []
        for row in template_LEDET_inputs_sheet:
            name, description = row[0], row[2]

            # If the row defines an empty row, or a title row, leave it unchanged
            if name == None:
                LEDET_inputs_sheet.append(row)  # Leave the row unchanged
                continue  # stop treating this row: go to the next row

            # Get value of the current parameter
            value = self.builder_ledet.getAttribute(name_sheet_inputs, name)

            # Skip optional variables if they have 0 elements, or if they are all NaN
            if name in optional_variables_input_sheet:
                if isinstance(value, list) and len(value) == 0:
                    if verbose: print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    continue
                elif isinstance(value, np.ndarray) and value.shape[0] == 0:
                    if verbose: print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    continue
                elif value is None:
                    if verbose: print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    continue
                elif isinstance(value, np.ndarray) and np.all(np.isnan(value)):
                    if verbose: print('Variable {} is optional and has only NaN elements, hence it will be skipped.'.format(name))
                    continue

            # Assign value to the variable sheet
            LEDET_inputs_sheet.append([name, value, description])

        ### Options sheet
        name_sheet_options = 'Options'  # This defines the name of the sheet and also of the variable group
        LEDET_options_sheet = []
        for row in template_LEDET_options_sheet:
            name, description = row[0], row[2]

            # If the row defines an empty row, or a title row, leave it unchanged
            if name == None:
                LEDET_options_sheet.append(row)  # Leave the row unchanged
                continue  # stop treating this row: go to the next row

            # Get value of the current parameter
            value = self.builder_ledet.getAttribute(name_sheet_options, name)

            # Skip optional variables if they have 0 elements
            if name in optional_variables_output_sheet:
                if isinstance(value, list) and len(value) == 0:
                    if verbose: print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    continue
                elif isinstance(value, np.ndarray) and value.shape[0] == 0:
                    if verbose: print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    continue
                elif value is None:
                    if verbose: print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    continue

            # Assign value to the variable sheet
            LEDET_options_sheet.append([name, value, description])

        ### Plots sheet
        name_sheet_plots = 'Plots'  # This defines the name of the sheet and also of the variable group
        LEDET_plots_sheet = []
        for row in template_LEDET_plots_sheet:
            name, description = row[0], row[2]

            # If the row defines an empty row, or a title row, leave it unchanged
            if name == None:
                LEDET_plots_sheet.append(row)  # Leave the row unchanged
                continue  # stop treating this row: go to the next row

            # Assign value to the variable sheet
            value = self.builder_ledet.getAttribute(name_sheet_plots, name)
            LEDET_plots_sheet.append([name, value, description])

        ### Variables sheet
        name_sheet_variables = 'Variables'  # This defines the name of the sheet and also of the variable group
        LEDET_variables_sheet = []
        for row in template_LEDET_variables_sheet:
            name, description = row[0], row[2]

            # If the row defines an empty row, or a title row, leave it unchanged
            if name == None:
                LEDET_variables_sheet.append(row)  # Leave the row unchanged
                continue  # stop treating this row: go to the next row

            # Assign value to the variable sheet
            value = self.builder_ledet.getAttribute(name_sheet_variables, name)
            LEDET_variables_sheet.append([name, value, description])

        # Write LEDET Excel input file
        write2Excel(name_file=full_path_file_name,
                    name_sheets=[name_sheet_inputs, name_sheet_options, name_sheet_plots, name_sheet_variables],
                    listOf_variables_values_descriptions=[LEDET_inputs_sheet, LEDET_options_sheet, LEDET_plots_sheet, LEDET_variables_sheet],
                    verbose=verbose)


    def write2json(self, full_path_file_name: str, verbose: bool = True, SkipConsistencyCheck: bool = False):
        '''
            *** Function that writes a LEDET input file as a .json file ***

            Function to write a LEDET input file containing "Inputs", "Options", "Plots", and "Variables" variables

            :param full_path_file_name: String defining the name of the LEDET input file to be written
            :type full_path_file_name: string
            :param verbose: flag that determines whether the output are printed
            :type verbose: bool
            :param SkipConsistencyCheck: flag that determines, whether the parameters shall be checked for consistency or not [False = Apply checks, True = Skip checks]
            :type SkipConsistencyCheck: bool

            :return: None
        '''

        # Unpack variables
        builder_ledet = self.builder_ledet

        # Consistency check
        if not SkipConsistencyCheck:
            if self._consistencyCheckLEDET():
                print("Variables are not consistent! Writing aborted - ", full_path_file_name)
                return
            else:
                if verbose: print("Preliminary consistency check was successful! - ", full_path_file_name)
        else:
            print("Skipped consistency checks.")

        # Define optional variables, which will be written only if present in the dataclass
        optional_variables_input_sheet = _getOptionalVariables_input()
        optional_variables_output_sheet = _getOptionalVariables_output()

        # Write LEDET data into a dictionary
        ledet_data_dict = {
            **builder_ledet.Inputs.__dict__,
            **builder_ledet.Options.__dict__,
            **builder_ledet.Plots.__dict__,
            **builder_ledet.Variables.__dict__}  # to add program-specific variables

        # Skip optional variables if they have 0 elements, or if they are all NaN
        for name, value in ledet_data_dict.copy().items():
            if (name in optional_variables_input_sheet) or (name in optional_variables_output_sheet):
                if isinstance(value, list) and len(value) == 0:
                    if verbose: print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    del ledet_data_dict[name]
                elif isinstance(value, np.ndarray) and value.shape[0] == 0:
                    if verbose: print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    del ledet_data_dict[name]
                elif value is None:
                    if verbose: print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    del ledet_data_dict[name]
                elif isinstance(value, np.ndarray) and np.all(np.isnan(value)):
                    if verbose: print('Variable {} is optional and has only NaN elements, hence it will be skipped.'.format(name))
                    del ledet_data_dict[name]

        # Write output .json file
        with open(full_path_file_name, 'w') as outfile:
            json.dump(ledet_data_dict, outfile, cls=NumpyEncoder, indent=4)

        # Display time stamp
        currentDT = datetime.datetime.now()
        if verbose:
            print(' ')
            print('Time stamp: ' + str(currentDT))
            print('New file ' + full_path_file_name + ' generated.')
        return


    def write2yaml(self, full_path_file_name: str, verbose: bool = True, SkipConsistencyCheck: bool = False):
        '''
            *** Function that writes a LEDET input file as a .yaml file ***

            Function to write a LEDET input file containing "Inputs", "Options", "Plots", and "Variables" variables

            :param full_path_file_name: String defining the name of the LEDET input file to be written
            :type full_path_file_name: string
            :param verbose: flag that determines whether the output are printed
            :type verbose: bool
            :param SkipConsistencyCheck: flag that determines, whether the parameters shall be checked for consistency or not [False = Apply checks, True = Skip checks]
            :type SkipConsistencyCheck: bool

            :return: None
        '''

        # Unpack variables
        builder_ledet = self.builder_ledet

        # Consistency check
        if not SkipConsistencyCheck:
            if self._consistencyCheckLEDET():
                print("Variables are not consistent! Writing aborted - ", full_path_file_name)
                return
            else:
                if verbose: print("Preliminary consistency check was successful! - ", full_path_file_name)
        else:
            print("Skipped consistency checks.")

        # Define optional variables, which will be written only if present in the dataclass
        optional_variables_input_sheet = _getOptionalVariables_input()
        optional_variables_output_sheet = _getOptionalVariables_output()

        # Write LEDET data into a dictionary
        ledet_data_dict = {
            **builder_ledet.Inputs.__dict__,
            **builder_ledet.Options.__dict__,
            **builder_ledet.Plots.__dict__,
            **builder_ledet.Variables.__dict__}  # to add program-specific variables

        # Skip optional variables if they have 0 elements, or if they are all NaN
        for name, value in ledet_data_dict.copy().items():
            if (name in optional_variables_input_sheet) or (name in optional_variables_output_sheet):
                if isinstance(value, list) and len(value) == 0:
                    if verbose: print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    del ledet_data_dict[name]
                elif isinstance(value, np.ndarray) and value.shape[0] == 0:
                    if verbose: print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    del ledet_data_dict[name]
                elif value is None:
                    if verbose: print('Variable {} is optional and has 0 elements, hence it will be skipped.'.format(name))
                    del ledet_data_dict[name]
                elif isinstance(value, np.ndarray) and np.all(np.isnan(value)):
                    if verbose: print('Variable {} is optional and has only NaN elements, hence it will be skipped.'.format(name))
                    del ledet_data_dict[name]

        # Write output .yaml file
        dict_to_yaml(ledet_data_dict, full_path_file_name, list_exceptions=[])

        # Display time stamp
        currentDT = datetime.datetime.now()
        if verbose:
            print(' ')
            print('Time stamp: ' + str(currentDT))
            print('New file ' + full_path_file_name + ' generated.')
        return


    def read_from_json(self, file_name: str, verbose: bool = True):
        '''
            *** Function that reads a .json file defining a LEDET input file and imports it in a BuilderLEDET object ***

            :param file_name: Name of the file to read
            :type file_name: str
            :param verbose: Flag indicating whether additional information should be displayed
            :type verbose: str

            :return: None
        '''

        # Read file as a dictionary
        with open(file_name) as json_file:
            dict_data = json.load(json_file)

        # Assign dictionary keys to the four LEDET dataclasses
        self._assign_dict_to_LEDET_dataclasses(dict_data)

        if verbose:
            print(f'File {file_name} successfully read.')


    def read_from_yaml(self, file_name: str, verbose: bool = True):
        '''
            *** Function that reads a .yaml file defining a LEDET input file and imports it in a BuilderLEDET object ***

            :param file_name: Name of the file to read
            :type file_name: str
            :param verbose: Flag indicating whether additional information should be displayed
            :type verbose: str

            :return: None
        '''

        # Read file as a dictionary
        yaml = ruamel.yaml.YAML(typ='safe', pure=True)
        with open(file_name) as yaml_file:
            dict_data = yaml.load(yaml_file)

        # Assign dictionary keys to the four LEDET dataclasses
        self._assign_dict_to_LEDET_dataclasses(dict_data)

        if verbose:
            print(f'File {file_name} successfully read.')


    #######################  Helper methods - START  #######################
    def _expand_scalar_to_array(self, verbose: bool = False):
        ''' Method adds as many elements as the number of half-turns to selected variables '''
        # TODO: Method can be made more general

        list_variables_to_expand = [
            'sim3D_f_cooling_down',
            'sim3D_f_cooling_up',
            'sim3D_f_cooling_left',
            'sim3D_f_cooling_right',
            ]

        for var_name in list_variables_to_expand:
            var_value = rgetattr(self.builder_ledet.Inputs, var_name)
            if isinstance(var_value, (int, float)):
                if verbose: print(f'{var_name} is a scalar. The same value will be written for all half-turns.')
            elif isinstance(var_value, (np.ndarray, list)) and len(var_value) == 1:
                if verbose: print(f'{var_name} is an array or a list of one element. The same value will be written for all half-turns.')
            else:
                continue  # do nothing
            # If the code reaches this line, assign the scalar value to all half-turns
            rsetattr(self.builder_ledet.Inputs, var_name, np.array([var_value] * int(np.sum(self.builder_ledet.Inputs.nT))))


    def _assign_dict_to_LEDET_dataclasses(self, dict_data: Dict):
        '''
        Assign the keys of a flat dictionary to the four LEDET dataclasses of a BuilderLEDET object
        :param dict_data:
        :return:
        '''
        for key, value in dict_data.items():
            if key in LEDETInputs.__annotations__:
                self.builder_ledet.setAttribute(self.builder_ledet.Inputs, key, value)
            if key in LEDETOptions.__annotations__:
                self.builder_ledet.setAttribute(self.builder_ledet.Options, key, value)
            if key in LEDETPlots.__annotations__:
                self.builder_ledet.setAttribute(self.builder_ledet.Plots, key, value)
            if key in LEDETVariables.__annotations__:
                self.builder_ledet.setAttribute(self.builder_ledet.Variables, key, value)

    #######################  Helper methods - END  #######################


#######################  Methods for consistency checks of LEDET input files - START  #######################
# TODO: Update consistency checks of LEDET input files to the latest LEDET features
    def __getNumberOfCoilSections(self):
        '''
            **Consistency check of LEDET Inputs - Helper function**

            Returns the number of CoilSections from Mutual-inductance matrix

            :return: int
        '''
        k = self.Inputs.M_m
        if k.shape == (1,): return k.shape[0]
        try:
            if k.shape[0] != k.shape[1]: print("M_m is not square")
        except:
            print("M_m is not square")
            return -1
        k2 = max(self.Inputs.GroupToCoilSection)
        if k.shape[0] != k2:
            print('M_m matrix does have size: ',k.shape[0], ' but you assign a Coil-Section: ',k2)
            return -1
        return k.shape[0]

    def __checkM_InductanceBlock_m(self, Turns):
        '''
            **Consistency check of LEDET Inputs - Inductance Matrix **

            Check if Inductance matrix is squared. Issues a warning if it is not the size of the number of turns. Returns result as bool

            :param Turns: Number of turns of the LEDET object
            :type Turns: int
            :return: bool
        '''
        if type(self.Inputs.M_InductanceBlock_m) != np.ndarray:
            k = np.array(self.Inputs.M_InductanceBlock_m)
        else:
            k = self.Inputs.M_InductanceBlock_m
        ## Account for the option to set the matrix to 0
        if k.shape == (1,):
            return True
        try:
            if k.shape[0] == k.shape[1]:
                if k.shape[0] != Turns:
                    print("M_InductanceBlock_m is squared, but its size unequal to the number of turns")
                return True
        except:
            print("M_InductanceBlock_m is not correct!")
            return False
        print("M_InductanceBlock_m is not correct!")
        return False

    def __checkHeFraction(self, Groups):
        '''
            **Consistency check of LEDET Inputs - Helium options check **

            Check if Helium options are both set and have the correct and same size, returns bool of result

            :param Groups: Number of groups of the LEDET object
            :type Groups: int
            :return: bool
        '''
        k = self.Inputs.overwrite_f_externalVoids_inGroup
        k2 = self.Inputs.overwrite_f_internalVoids_inGroup
        if len(k) > 0:
            if len(k) != len(k2):
                print("Helium section was set but is corrupted.")
                return False
            if len(k) != Groups:
                print("Helium section was set but is wrong length.")
                return False
        elif len(k2) > 0:
            print("Helium section was set but is corrupted.")
            return False
        return True

    def __checkMonotony(self, arr, Invert=False):
        '''
            **Consistency check of LEDET Inputs - Monotony check **

            Check if given array is monotone or not. Returns bool of result.

            :param arr: Given array to be checked
            :type arr: np.ndarray()
            :param Invert: flag that determines which direction the array should be interpreted [True= from the last towards the first, False= from the first towards the last]
            :type Invert: bool
            :return: bool
        '''
        if Invert:
            arr = np.flip(arr)
        b = all(x <= y for x, y in zip(arr, arr[1:]))
        return b

    def __checkTimes(self):
        '''
            **Consistency check of LEDET Inputs - check Times **

            Function that checks if times in LEDET are all set accordingly, otherwise adjusts the time_vector
            :return: none
        '''

        ### Check start of time_vector
        # obtain all times from LEDET object
        try: t1 = min(self.Inputs.tQuench)
        except: t1 = min(np.array([self.Inputs.tQuench]))
        t2 = self.Inputs.t_PC_LUT[0]
        try: t3 = min(self.Inputs.tStartQuench)
        except: t3 = min(np.array([self.Inputs.tStartQuench]))
        t4 = self.Options.time_vector_params[0]
        try:  t5 = min(self.Inputs.tQH)
        except: t5 = min(np.array([self.Inputs.tQH]))
        t6 = self.Inputs.tCLIQ
        t7 = self.Inputs.tEE

        # Check if times are all after the beginning of the simulation, otherwise set the beginning to the earliest time used
        if any(x < t4 for x in [t1,t2,t3, t5, t6, t7]):
            print("You're using a time, that is before the start of the simulation. Corrected Time-Vector.")
            self.Options.time_vector_params[0] = np.min([t1, t2, t3, t5, t6, t7])-0.01
            self.Inputs.t_PC_LUT[0] = np.min([t1, t2, t3, t5, t6, t7])-0.01
            self.Inputs.tQuench = np.zeros((len(self.Inputs.tQuench),))+np.min([t1, t2, t3, t5, t6, t7])

        ### Check end of time vector
        # obtain all times from LEDET object [if times are above > 999, they are interpreted as not set]
        try: t1 = max(self.Inputs.tQuench)
        except: t1 = max(np.array([self.Inputs.tQuench]))
        if t1 >= 999: t1 = 0
        t2 = self.Inputs.t_PC_LUT[-1]
        if t2 >= 999: t2 = 0
        t3 = self.Inputs.tStartQuench
        try: t3 = [0 if x>=999 else x for x in t3]
        except:
            t3 = [t3]
            t3 = [0 if x >= 999 else x for x in t3]
        t3 = max(t3)
        t4 = self.Options.time_vector_params[-1]
        t5 = self.Inputs.tQH
        try: t5 = [0 if x>=999 else x for x in t5]
        except:
            t5 = [t5]
            t5 = t5 = [0 if x>=999 else x for x in t5]
        t5 = max(t5)
        t6 = self.Inputs.tCLIQ
        if t6 >= 999: t6 = 0
        t7 = self.Inputs.tEE
        if t7 >= 999: t7 = 0

        # Check if times are all before the end of the simulation, otherwise extend the time_vector
        if any(x > t4 for x in [t1, t2, t3, t5, t6, t7]):
            print("You're using a time, that is after the end of the simulation. Corrected Time-Vector.")
            self.Options.time_vector_params[-1] = np.max([t1, t2, t3, t5, t6, t7]) + 1
            self.Inputs.t_PC_LUT = np.append(self.Inputs.t_PC_LUT, np.max([t1, t2, t3, t5, t6, t7]) + 1)
            self.Inputs.I_PC_LUT = np.append(self.Inputs.I_PC_LUT, self.Inputs.I_PC_LUT[-1])
        return 1

    def __checkPersistentCurrents(self):
        if len(self.Inputs.df_inGroup)==1:
            if self.Options.flag_persistentCurrents == 0: return True
            else:
                print('Persistent current parameters flag is set, but parameters not. I set the flag to 0.')
                self.Options.flag_persistentCurrents = 0
                return True
        else:
            if self.Options.flag_persistentCurrents==0:
                print('Persistent current parameters are set but flag is not. Continuing.')
            maxFit = np.max(self.Inputs.selectedFit_inGroup)
            shp = self.Inputs.fitParameters_inGroup.shape
            if maxFit == 1:
                if not shp[0]>0 or not shp[1]==len(self.Inputs.nT):
                    print('You selected constant Jc, but fitParameters are not set. Abort.')
                    return False
                else: return True
            elif maxFit == 2:
                if not shp[0]>1 or not shp[1]==len(self.Inputs.nT):
                    print('You selected Botturas fit, but not enough fit parameters provided. Abort.')
                    return False
                else:
                    return True
            elif maxFit == 3:
                if not shp[0] > 7 or not shp[1] == len(self.Inputs.nT):
                    print('You selected CUDI fit, but not enough fit parameters provided. Abort.')
                    return False
                else:
                    return True
            elif maxFit == 4:
                if not shp[0] > 2 or not shp[1] == len(self.Inputs.nT):
                    print('You selected Summers fit, but not enough fit parameters provided. Abort.')
                    return False
                else:
                    return True
            else:
                print('Unknown fit. Please check! Abort.')
                return False

    def __checkThermalConnections(self):
        if len(self.Inputs.iContactAlongHeight_From) == 0 or len(self.Inputs.iContactAlongHeight_To) == 0:
            self.Inputs.iContactAlongHeight_From = np.array([1])
            self.Inputs.iContactAlongHeight_To = np.array([1])
            print('No thermal connections in height directions set. I added at least 1.')
        if len(self.Inputs.iContactAlongWidth_From) == 0 or len(self.Inputs.iContactAlongWidth_To) == 0:
            self.Inputs.iContactAlongWidth_From = np.array([1])
            self.Inputs.iContactAlongWidth_To = np.array([1])
            print('No thermal connections in width directions set. I added at least 1.')

    def _consistencyCheckLEDET(self):
        '''
            **Consistency check of LEDET Inputs - Main function **

            Function applies different consistency checks on LEDET inputs to see if the values are set properly
            Applied checks:
                - Length checks [checking Inputs that require the same size]
                - checkM_InductanceBlock_m [checking if inductance matrix is squared]
                - checkHeFraction [checking if both Helium options are set]
                - checkMonotony [check Inputs that require monotony in themselves]
                - checkTimes [check if TimeVector fits to other times in the Inputs]

            :return Break: flag, showing if Inputs are consistent or not
            :type break: bool
        '''
        # Define groups that require the same size, number in each list contains the row-number of the attribute
        ## 0 Single - 1 CoilSections - 2 Groups - 3 Half-Turns - 4 doesn't matter - 5 iContactAlongWidth - 6 iContactAlongHeight - 7 vQlength
        ## 8 Quench Heater, 9 QH_QuenchToFrom, 10 CLIQ, 11 Persistent currents
        VarsSameInput = [['T00','l_magnet','I00','R_circuit','R_crowbar','Ud_crowbar','t_PC', 't_EE', 'R_EE_triggered', 'sim3D_uThreshold',
                          'sim3D_f_cooling_down','sim3D_f_cooling_up','sim3D_f_cooling_left','sim3D_f_cooling_right','sim3D_fExToIns',
                          'sim3D_fExUD','sim3D_fExLR','sim3D_min_ds_coarse','sim3D_min_ds_fine','sim3D_min_nodesPerStraightPart',
                          'sim3D_min_nodesPerEndsPart','sim3D_Tpulse_sPosition','sim3D_Tpulse_peakT',
                          'sim3D_Tpulse_width','sim3D_durationGIF','sim3D_flag_saveFigures','sim3D_flag_saveGIF',
                          'sim3D_flag_VisualizeGeometry3D','sim3D_flag_SaveGeometry3D'],
                         ['M_m', 'directionCurrentCLIQ', 'tQuench', 'initialQuenchTemp'],
                         ['GroupToCoilSection','polarities_inGroup','nT','nStrands_inGroup','l_mag_inGroup','ds_inGroup','f_SC_strand_inGroup',\
                         'f_ro_eff_inGroup','Lp_f_inGroup','RRR_Cu_inGroup','SCtype_inGroup','STtype_inGroup','insulationType_inGroup', \
                         'internalVoidsType_inGroup', 'externalVoidsType_inGroup', 'wBare_inGroup', 'hBare_inGroup','wIns_inGroup','hIns_inGroup',\
                         'Lp_s_inGroup', 'R_c_inGroup', 'Tc0_NbTi_ht_inGroup', 'Bc2_NbTi_ht_inGroup', 'c1_Ic_NbTi_inGroup','c2_Ic_NbTi_inGroup',\
                         'Tc0_Nb3Sn_inGroup','Bc2_Nb3Sn_inGroup','Jc_Nb3Sn0_inGroup','el_order_half_turns'],
                        ['el_order_half_turns', 'alphasDEG', 'rotation_block', 'mirror_block', 'mirrorY_block','HalfTurnToInductanceBlock'],
                        ['fL_I', 'fL_L', 'overwrite_f_internalVoids_inGroup', 'overwrite_f_externalVoids_inGroup','t_PC_LUT','I_PC_LUT', 'sim3D_idxFinerMeshHalfTurn'],
                        ['iContactAlongWidth_From', 'iContactAlongWidth_To'],
                        ['iContactAlongHeight_From', 'iContactAlongHeight_To'],
                        ['iStartQuench', 'tStartQuench', 'lengthHotSpot_iStartQuench', 'fScaling_vQ_iStartQuench'],
                        ['tQH', 'U0_QH', 'C_QH', 'R_warm_QH', 'w_QH', 'h_QH', 's_ins_QH', 'type_ins_QH', 's_ins_QH_He', 'type_ins_QH_He',\
                         'l_QH', 'f_QH'],
                        ['iQH_toHalfTurn_From','iQH_toHalfTurn_To'],
                        ['tCLIQ','nCLIQ','U0', 'C', 'Rcapa'],
                        ['df_inGroup', 'selectedFit_inGroup', 'fitParameters_inGroup']]


        slicesSameInput = []
        for i in range(len(VarsSameInput)):
            slicesSameInput.append([])

        counter = 1
        for l in self.Inputs.__annotations__:
            for i in range(len(VarsSameInput)):
                try:
                    _ = VarsSameInput[i].index(l)
                    slicesSameInput[i].append(counter-1)
                except:
                    pass
            counter = counter + 1

        # Acquire representative sizes for the defined groups
        lengthInputs = len(self.Inputs.__annotations__)
        sizeInputs = np.zeros((lengthInputs,1))
        sizeInputs[slicesSameInput[0]] = 1 #single Valued
        sizeInputs[slicesSameInput[1]] = self.__getNumberOfCoilSections() #Number of CoilSections

        if sizeInputs[slicesSameInput[1][0]] == -1:
            print("M_m or Number of Coilsections is corrupted. please check.")
            return True
        sizeInputs[slicesSameInput[2]] = len(self.Inputs.nT) #Number of Groups
        sizeInputs[slicesSameInput[3]] = sum(self.Inputs.nT) #Number of Turns
        sizeInputs[slicesSameInput[4]] = 0 #Unchecked
        sizeInputs[slicesSameInput[5]] = len(self.Inputs.iContactAlongWidth_From)
        sizeInputs[slicesSameInput[6]] = len(self.Inputs.iContactAlongHeight_From)
        sizeInputs[slicesSameInput[7]] = len(self.Inputs.iStartQuench)
        sizeInputs[slicesSameInput[8]] = len(self.Inputs.tQH)
        sizeInputs[slicesSameInput[9]] = len(self.Inputs.iQH_toHalfTurn_From)
        try:
            sizeInputs[slicesSameInput[10]] = len(self.Inputs.tCLIQ)
        except:
            sizeInputs[slicesSameInput[10]] = 1
        if len(self.Inputs.df_inGroup)>1:
            sizeInputs[slicesSameInput[11]] = len(self.Inputs.nT)
        else:
            sizeInputs[slicesSameInput[11]] = len(self.Inputs.df_inGroup)

        # Checks for types and lengths/sizes
        Count = 0
        Break = 0
        for k in self.Inputs.__annotations__:
            if sizeInputs[Count] == 0:
                Count = Count + 1
                continue
            cC = self.getAttribute(self.Inputs, k)
            if type(cC) == list:
                if not len(cC)==sizeInputs[Count]:
                    print("The variable ", k, " does not have the correct size, should be", sizeInputs[Count]," but is ",len(cC),"! Please check.")
                    Break = 1
            elif type(cC) == np.ndarray:
                if len(cC.shape)==1:
                    if not len(cC)==sizeInputs[Count]:
                        print("The variable ", k, " does not have the correct size, should be", sizeInputs[Count]," but is ",len(cC),"! Please check.")
                        Break = 1
                else:
                    if not cC.shape[1]==sizeInputs[Count]:
                        print("The variable ", k, " does not have the correct size, should be", sizeInputs[Count]," but is ",cC.shape[1],"! Please check.")
                        Break = 1
            elif type(cC) == float or type(cC) == int or type(cC) == np.float64:
                if not sizeInputs[Count]== 1:
                    print("The variable ", k, " does not have the correct size, should be", sizeInputs[Count]," but is", type(cC),". Please check.")
                    Break = 1
            else:
                print("Variable ", k, " has the wrong data-type set! Please check.")
                Break = 1
            Count = Count + 1

        ## Remaining checks in functions
        if not self.__checkHeFraction(len(self.Inputs.nT)):
            Break = 1
        if not self.__checkMonotony(self.Inputs.t_PC_LUT):
            print("t_PC_LUT is not monotonic")
            Break = 1
        if not self.__checkMonotony(self.Inputs.fL_I):
            print("fL_I is not monotonic")
            Break = 1
        if not self.__checkTimes():
            Break = 1
        if not self.__checkPersistentCurrents():
            Break = 1
        if not self.__checkM_InductanceBlock_m(int(sum(self.Inputs.nT)/2)):
            Break = 1
        return Break
#######################  Methods for consistency checks of LEDET input files - END  #######################


#######################  Helper functions - START  #######################
def _getOptionalVariables_input():
    # Define optional variables, which will be written only if present in the dataclass
    optional_variables_input_sheet = [
        'alpha_Nb3Sn_inGroup',
        'f_scaling_Jc_BSCCO2212_inGroup',
        'overwrite_f_internalVoids_inGroup', 'overwrite_f_externalVoids_inGroup',
        'f_RRR1_Cu_inGroup', 'f_RRR2_Cu_inGroup', 'f_RRR3_Cu_inGroup',
        'RRR1_Cu_inGroup', 'RRR2_Cu_inGroup', 'RRR3_Cu_inGroup',
        'R_EE_power',
        ]
    return optional_variables_input_sheet


def _getOptionalVariables_output():
    optional_variables_output_sheet = [
        'fieldMapNumber',
        'selfMutualInductanceFileNumber',
        'flag_calculateMagneticField',
        'flag_controlInductiveVoltages',
        'flag_controlMagneticField',
        'flag_controlBoundaryTemperatures',
        ]
    return optional_variables_output_sheet


def CompareLEDETParameters(fileA: str, fileB: str, max_relative_error: float=1E-5, verbose=False):
    '''
        Compare all the variables imported from two LEDET Excel input files
        Returns True if the two files contain LEDET parameters that differ by less than a certain relative error. Returns False otherwise.
    '''

    pl_a = ParserLEDET(BuilderLEDET(flag_build=False))
    pl_a.readFromExcel(fileA, verbose=False)
    pl_b = ParserLEDET(BuilderLEDET(flag_build=False))
    pl_b.readFromExcel(fileB, verbose=False)
    print("Starting Comparison of A: ({}) and B: ({})".format(fileA, fileB))

    Diff = check_for_differences(pl_a, pl_b, max_relative_error, verbose=verbose)

    if Diff == False:
        print(f'Files {fileA} and {fileB} are equal: they contain LEDET parameters that differ by a relative error lower than {max_relative_error}.')
        return True
    else:
        return False


def check_for_differences(pl_a, pl_b, max_relative_error, verbose):
    Diff = False

    ## Check Inputs
    for attribute in pl_a.builder_ledet.Inputs.__annotations__:
        last_diff = compare_two_parameters(pl_a.builder_ledet.getAttribute("Inputs", attribute),
                                      pl_b.builder_ledet.getAttribute("Inputs", attribute),
                                      attribute, max_relative_error, flag_difference_messages=True, verbose=verbose)
        if last_diff: Diff = True

    ## Check Options
    for attribute in pl_a.builder_ledet.Options.__annotations__:
        last_diff = compare_two_parameters(pl_a.builder_ledet.getAttribute("Options", attribute),
                                      pl_b.builder_ledet.getAttribute("Options", attribute),
                                      attribute, max_relative_error, flag_difference_messages=True, verbose=verbose)
        if last_diff: Diff = True

    ## Check Plots
    for attribute in pl_a.builder_ledet.Plots.__annotations__:
        last_diff = compare_two_parameters(pl_a.builder_ledet.getAttribute("Plots", attribute),
                                      pl_b.builder_ledet.getAttribute("Plots", attribute),
                                      attribute, max_relative_error, flag_difference_messages=True, verbose=verbose)
        if last_diff: Diff = True

    ## Check Variables
    for attribute in pl_a.builder_ledet.Variables.__annotations__:
        last_diff = compare_two_parameters(pl_a.builder_ledet.getAttribute("Variables", attribute),
                                      pl_b.builder_ledet.getAttribute("Variables", attribute),
                                      attribute, max_relative_error, flag_difference_messages=True, verbose=verbose)
        if last_diff: Diff = True

    return Diff



def copy_map2d(magnet_name: str, map2d_file_name: str, output_path: str, flagIron: int, flagSelfField: int,
               suffix: str = '', verbose: bool = False):
    ''' rename and copy '''

    # Checks if path exists
    if output_path != '' and not os.path.isdir(output_path):
        print("Output folder {} does not exist. Making it now".format(output_path))
        Path(output_path).mkdir(parents=True)

    ### new naming of file (depending on magnet properties)

    if flagIron == 1:
        suffix_iron = "_WithIron"
    else:
        suffix_iron = "_NoIron"

    if flagSelfField == 1:
        suffix_self = "_WithSelfField"
    else:
        suffix_self = "_NoSelfField"

    suffix_complete = suffix + suffix_iron + suffix_self
    file_name_output = magnet_name + suffix_complete + ".map2d"

    # Paths
    output_path_full = os.path.join(output_path, file_name_output)

    # Copy
    shutil.copy2(map2d_file_name, output_path_full)

    if verbose:
        print('File {} copied to {}.'.format(map2d_file_name, output_path_full))

    return file_name_output


def copy_modified_map2d_ribbon_cable(magnet_name: str, map2d_file_name: str, output_path: str, geometry_ribbon_cable,
                                     flagIron: int, flagSelfField: int, list_flag_ribbon: list, suffix: str = '', verbose: bool = False):
    ''' #TODO '''

    ### Check if output path exists

    if output_path != '' and not os.path.isdir(output_path):
        print("Output folder {} does not exist. Making it now".format(output_path))
        Path(output_path).mkdir(parents=True)

    ### naming of file (depending on magnet properties)

    if flagIron == 1:
        suffix_iron = "_WithIron"
    else:
        suffix_iron = "_NoIron"

    if flagSelfField == 1:
        suffix_self = "_WithSelfField"
    else:
        suffix_self = "_NoSelfField"

    suffix_complete = suffix + suffix_iron + suffix_self
    file_name_output = magnet_name + suffix_complete + ".map2d"

    # Get data of (old) ROXIE file
    content_ROXIE_file = open(map2d_file_name, "r").read().split("\n")

    # Parse to ParserMap2d
    NewMap2d = ParserMap2d.modify_map2d_ribbon_cable(map2d_file_name, geometry_ribbon_cable, list_flag_ribbon)
    # Multiply 3rd,4th and 7th column to not have SI-Units
    ParserMap2d.multiply_column_by_value(NewMap2d, 3, 1000)
    ParserMap2d.multiply_column_by_value(NewMap2d, 4, 1000)
    ParserMap2d.multiply_column_by_value(NewMap2d, 7, 1000000)
    # Create .map2d file from ParserMap2d_
    ParserMap2d.create_map2d_file(NewMap2d, header_line=content_ROXIE_file[0],
                                  output_file_location=Path(output_path, file_name_output),
                                  new_file_name=file_name_output,verbose=verbose)

    return file_name_output
#######################  Helper functions - END  #######################