import time
import pytest
import allure
import openpyxl
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
from frame_lustre import dropDowBox_Value
from selenium.webdriver.common.action_chains import ActionChains


@pytest.fixture(scope="session")
def driver():
    driver = webdriver.Chrome()
    yield driver
    driver.quit()


class WebDriverBase:
    def __init__(self):
        self.driver = webdriver.Chrome()

    # 打开页面

    def open(self, url):
        self.driver.get(url)

    # 关闭浏览器

    def close(self):
        self.driver.close()

    # 退出浏览器

    def quit(self):
        self.driver.quit()

    # 定位元素

    def locate(self, way, op_object):
        try:
            sth = self.driver.find_element(way, op_object)
            return sth
        except NoSuchElementException as e:
            print(f"Element not found: {op_object}")
        raise


class NavigationHandler:
    def __init__(self, webdriver_base):
        self.driver = webdriver_base.driver

    # 后退

    def back(self):
        self.driver.back()

    # 前进

    def forward(self):
        self.driver.forward()

    # 刷新

    def refresh(self):
        self.driver.refresh()


class ElementInteraction:
    def __init__(self, webdriver_base):
        self.driver = webdriver_base.driver

    def click(self, way, op_object):
        try:
            element = self.driver.find_element(way, op_object)
            element.click()
        except Exception as e:
            print(f"Failed to click on element: {op_object}")
            raise

    def enter_text(self, way, op_object, text):
        element = self.driver.find_element(way, op_object)
        element.clear()
        element.send_keys(text)

    def get_text(self, way, op_object):
        element = self.driver.find_element(way, op_object)
        return element.text

    def clear_text(self, way, op_object):
        self.driver.find_element(way, op_object).clear()


class WaitHandler:
    def __init__(self, webdriver_base):
        self.driver = webdriver_base.driver

    def wait_for_element(self, way, op_object, timeout=10):
        try:
            element = WebDriverWait(self.driver, timeout).until(
                EC.presence_of_element_located((way, op_object))
            )
            return element
        except TimeoutException:
            print(f"Timeout waiting for element: {op_object}")
            return None

    def wait_for_visibility(self, way, op_object, timeout=10):
        try:
            element = WebDriverWait(self.driver, timeout).until(
                EC.visibility_of_element_located((way, op_object))
            )
            return element
        except TimeoutException:
            print(f"Timeout waiting for element's visibility: {op_object}")
            return None


class AssertHandler:
    def __init__(self, webdriver_base):
        self.driver = webdriver_base.driver

    def assert_text(self, way, op_object, expected_text):
        try:
            element = self.driver.find_element(way, op_object)
            assert element.text == expected_text, f"Expected text '{expected_text}' did not match actual '{element.text}'."
            return True
        except AssertionError as e:
            print(f"Assertion Error: {e}")
            return False


class ScreenshotHandler:
    def __init__(self, webdriver_base):
        self.driver = webdriver_base.driver

    def take_screenshot(self, screenshot_name):
        try:
            allure.attach(self.driver.get_screenshot_as_png(), name=screenshot_name,
                          attachment_type=allure.attachment_type.PNG)
        except Exception as e:
            print(f"Failed to take screenshot: {e}")


class FileHandler:
    def __init__(self, webdriver_base):
        self.driver = webdriver_base.driver

    def upload_file(self, way, op_object, file_path):
        try:
            element = self.driver.find_element(way, op_object)
            element.send_keys(file_path)
        except Exception as e:
            print(f"Failed to upload file: {e}")

    def drag_and_drop_file(self, way, op_object, file_path):
        try:
            element = self.driver.find_element(way, op_object)
            action_chains = ActionChains(self.driver)
            action_chains.drag_and_drop(file_path, element).perform()
        except Exception as e:
            print(f"Failed to drag and drop file: {e}")


class DropdownHandler:
    def __init__(self, webdriver_base):
        self.driver = webdriver_base.driver

    def select_option_by_text(self, way, op_object, option_text):
        try:
            element = self.driver.find_element(way, op_object)
            select = Select(element)
            select.select_by_visible_text(option_text)
        except Exception as e:
            print(f"Failed to select option by text: {e}")


class ExcelHandler:
    @staticmethod
    def read_all_sheets(file_path):
        workbook = openpyxl.load_workbook(file_path)
        all_data = {}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            field_names = [cell.value for cell in sheet[1]]

            sheet_data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(field_names, row))
                sheet_data.append(row_data)

            all_data[sheet_name] = sheet_data

        workbook.close()
        return all_data


class PageDataHandler:
    def __init__(self, webdriver_base):
        self.driver = webdriver_base.driver

    def get_data_from_table_column(self, way, op_object, column_index):
        try:
            table = self.driver.find_element(way, op_object)
            rows = table.find_elements(By.TAG_NAME, "tr")
            data_list = []

            if 0 <= column_index < len(rows[0].find_elements(By.TAG_NAME, "td")):
                for row in rows:
                    columns = row.find_elements(By.TAG_NAME, "td")
                    specific_column_value = columns[column_index].text
                    data_list.append(specific_column_value)

            return data_list
        except Exception as e:
            print(f"Failed to get data from table column: {e}")
