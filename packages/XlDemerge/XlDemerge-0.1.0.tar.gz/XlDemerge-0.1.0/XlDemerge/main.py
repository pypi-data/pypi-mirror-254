import openpyxl
import pandas
import numpy

def cell_unmerge(file_path, sheet_name):
    # Load workbook and select worksheet
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook[sheet_name]

    # Get merged cell ranges
    merged_cells = list(worksheet.merged_cells.ranges)

    # Unmerge cells and fill with values
    for merged_cell_range in merged_cells:
        min_col, min_row, max_col, max_row = merged_cell_range.bounds
        top_left_cell_value = worksheet.cell(row=min_row, column=min_col).value
        worksheet.unmerge_cells(str(merged_cell_range))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                worksheet.cell(row=row, column=col, value=top_left_cell_value)

    # Convert worksheet data to pandas DataFrame
    data = worksheet.values
    columns = next(data)
    df = pandas.DataFrame(data, columns=columns)

    # Replace 'None' with 'np.nan'
    df = df.where(pandas.notnull(df), numpy.nan)

    # Remove rows where all values are NaN
    df = df.dropna(how='all')

    return df

def rows_unmerge(file_path, sheet_name):
    # Load workbook and select worksheet
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook[sheet_name]

    # Get merged cell ranges that span across rows
    merged_cells = [merge for merge in worksheet.merged_cells.ranges if merge.min_row != merge.max_row]

    # Unmerge cells and fill with values
    for merged_range in merged_cells:
        top_left_cell_value = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
        worksheet.unmerge_cells(str(merged_range))
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            worksheet.cell(row=row, column=merged_range.min_col).value = top_left_cell_value

    # Convert worksheet data to pandas DataFrame
    data = worksheet.values
    columns = next(data)
    df = pandas.DataFrame(data, columns=columns)

    # Replace 'None' with 'np.nan'
    df = df.where(pandas.notnull(df), numpy.nan)

    # Remove rows where all values are NaN
    df = df.dropna(how='all')

    return df

def columns_unmerge(file_path, sheet_name):
    # Load workbook and select worksheet
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook[sheet_name]

    # Get merged cell ranges that span across columns
    merged_cells = [merge for merge in worksheet.merged_cells.ranges if merge.min_col != merge.max_col]

    # Unmerge cells and fill with values
    for merged_range in merged_cells:
        top_left_cell_value = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
        worksheet.unmerge_cells(str(merged_range))
        for col in range(merged_range.min_col, merged_range.max_col + 1):
            worksheet.cell(row=merged_range.min_row, column=col).value = top_left_cell_value

    # Convert worksheet data to pandas DataFrame
    data = worksheet.values
    columns = next(data)
    df = pandas.DataFrame(data, columns=columns)

    # Replace 'None' with 'np.nan'
    df = df.where(pandas.notnull(df), numpy.nan)
    
    # Remove rows where all values are NaN
    df = df.dropna(how='all')

    return df

def merged_cell_range(file_path, sheet_name):
    list_merged_cell = []

    # Load workbook and select worksheet
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook[sheet_name]

    # Get merged cell ranges
    merged_cells = worksheet.merged_cells.ranges

    # Append merged cell ranges to list
    for merged_cell in merged_cells:
        list_merged_cell.append(f'{merged_cell}')

    return list_merged_cell