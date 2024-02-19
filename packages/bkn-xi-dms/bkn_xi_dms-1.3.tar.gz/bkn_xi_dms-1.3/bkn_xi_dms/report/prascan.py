from bkn_xi_dms.utill.installed_application import *
import pandas as pd
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows


def create_report(json):
    return {
        "job_id": json["data"]["uniqueCode"],
        "qty": json["data"]["documentQty"],
        "tanggal":  json["data"]["createdDate"]
    }


def create_report_excel(list_data, output_path, sheet, color):
    list_data = pd.DataFrame(list_data)
    if os.path.isfile(output_path):
        workbook = openpyxl.load_workbook(output_path)
        if sheet in workbook.sheetnames:
            existing_data = pd.read_excel(output_path, sheet_name=sheet)
            updated_data = pd.concat([existing_data, list_data])
            worksheet = workbook[sheet]
            worksheet.sheet_properties.tabColor = color
            for row in dataframe_to_rows(updated_data, index=False, header=False):
                worksheet.append(row)
            workbook.save(output_path)
        else:
            worksheet = workbook.create_sheet(title=sheet)
            worksheet.sheet_properties.tabColor = color
            for row in dataframe_to_rows(list_data, index=False, header=True):
                worksheet.append(row)
        workbook.save(output_path)
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet
        worksheet.sheet_properties.tabColor = color
        for row in dataframe_to_rows(list_data, index=False, header=True):
            worksheet.append(row)
        workbook.save(output_path)
