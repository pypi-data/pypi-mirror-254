# vision_oslo_extension/excel_processing.py
import pandas as pd
import numpy as np
import os
# import openpyxl
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle
from openpyxl.formatting.rule import CellIsRule,FormulaRule
from openpyxl.chart import ScatterChart,Reference,Series
from openpyxl.drawing.text import CharacterProperties
from openpyxl.chart.shapes import GraphicalProperties

# import vision_oslo
from vision_oslo_extension import oslo_extraction
from vision_oslo_extension.shared_contents import SharedMethods


def main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step):
    
    simname = simname
    # Specify Excel file name
    excel_file = text_input + ".xlsx"

    if not SharedMethods.check_existing_file(excel_file):
        return False
    
    #option = "1" # Fully Restart, require oof,
    #option = "2" # Process only

    option = option_select # 1:Fully Restart, require oof, and list

    start = 7 # result start from row 7
    space = 5
    # time_increment = 5

    time_windows= ['1min','4min','5min','30min','60min','120min','180min']

    if option not in ["0","1","2"]:
        print("ERROR: Error in dc_tru.py. Please contact Support...")
        return False
    
    if option == "0":
        print("ERROR: Please select an option to proceed.")
        return False

    # start_cell = 'B11'
    # read data from start tab
    # start_df = pd.read_excel(writer,sheet_name = "Start")
    result = start_reading_process(simname,excel_file)
    if result == False:
        return False
    else:
        start_df,oslo_total,scenariolist = result

    # branch_list: branch format in XXXX/X format
    # branch_list_new: branch format in XXXX-X format this is because many application does not allow / charaters in naming
        

    # check if want to go throught the feeder check process
    if option == "1":
        # check essential files
        print('Checking essential files...')
        for out,value in scenariolist.items():
            folder = value[0]
            sim = value[1]
            if folder == None or sim == None:
                print("WARNING: Scenario List Information Not Complete. Please check.")
            else:
                sim = sim + "_RMSCurrent_Sum.csv"
                if not SharedMethods.folder_file_check(folder,sim):
                    print("ERROR: Check info above. Adjust Setting or Do extraction first")
                    return False
        
    if option == "2":
        print('ERROR: Option to be developed')
        return False
    
    # read all essential files: rating summary , result summary
    t_df_sum, r_df_sum = read_all_files(start_df,oslo_total,scenariolist,time_windows)

    # processing the information
    t_df_sum = result_update(start_df,oslo_total,scenariolist,time_windows,t_df_sum,r_df_sum)

    # check failure summary
    fail_df_sum = check_failure(t_df_sum,scenariolist)

    # substation group
    sub_df_sum = substation_group(fail_df_sum,scenariolist,time_windows)

    # save the data
    data_write_save(simname,excel_file,start,space,oslo_total,scenariolist,time_windows,t_df_sum,fail_df_sum)

    return True




# write data to excel
def data_write_save(simname,excel_file,start,space,oslo_total,scenariolist,time_windows,t_df_sum,fail_df_sum):

    # write data to excel
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a',if_sheet_exists='overlay') as writer:

        # get currrent workbook to do process
        wb = writer.book

        print("Generate Failure Result Page...")
        r_start = start + 2
        for index, dflist in enumerate(fail_df_sum):
            dflist.to_excel(writer, sheet_name="Result", index=False, startrow = r_start, startcol = 1)
            r_start = r_start+ dflist.shape[0] + space

        # start_df.to_excel(writer, sheet_name="Result", index=False, startrow = start, startcol = 1)
        # result_df.to_excel(writer, sheet_name="Result", index=False, startrow = start, startcol = start_df.shape[1] + 1)

        # # print("Generate Plot Page...")
        # print("Saving {}...".format(feeder))

        # saving individual feeder
        for index, dflist in enumerate(t_df_sum):

            print("Saving {} Result...".format(time_windows[index]))
            # emty columns
            # dflist.insert(dflist.columns.get_loc('I_angle')+1,'New_C_1', np.nan)

            #sumdataframe[index].insert(sumdataframe[index].columns.get_loc('I_angle')+1,'New_C_1', np.nan)

            #sumdataframe[index].to_excel(writer, sheet_name=branch_list_new[index], index=False, startrow = 0)

            dflist.to_excel(writer, sheet_name=time_windows[index], index=False, startrow = start-1)

        
        # # Calculate the Excel range for post-processing
        range_list = get_result_range(start,space,oslo_total,scenariolist,fail_df_sum)

        # # do the plot
        # total = len(d4_df)
        # curve_plot(wb,feeder,range_list,result_df,plot_start,total)
        
        # # table formatting
        table_formatting(simname,wb,range_list,time_windows,scenariolist)

        print("Saving Excel File...")

    return

# read the start tab and collect informaiton
def start_reading_process(simname, excel_file):
    
    # File check
    print("Check Excel and Deleting Existing Contents ...")

    try:
        wb = load_workbook(excel_file)
        # Delete all sheets except 'Start'
        for sheet_name in wb.sheetnames:
            if sheet_name == 'Start':
                result = check_reading_frame(wb[sheet_name]) # list, dictionary type
                if result == False:
                    return False
                else:
                    data_start_row,data_end_row,oslo_total,scenariolist = result
                
                columns = ['Substation Name', 'OSLO', 'Outage Scenario', 'TRU Rating (MW)', 'TRU Type', \
                           '1min', '4min', '5min', '30min','60min','120min','180min', \
                            '1min_out', '4min_out', '5min_out', '30min_out','60min_out','120min_out','180min_out']
                start_df = pd.read_excel(excel_file,sheet_name = 'Start',header = 0,usecols=range(19), skiprows=data_start_row-1, nrows=oslo_total+1, names = columns)
                oslolist = start_df.iloc[:, 1].tolist()
                outlist =  start_df.iloc[:, 2].tolist()

                flag = False
                # check oslo list
                for oslo in oslolist:
                    if pd.isna(oslo):
                        print('Warning: Substation {} is not assigned with OSLO ID'.format(start_df.iloc[index,0]))
                        flag = True
                
                # Check if all numbers in outlist are in the keys of scearniolist
                for index, out in enumerate(outlist):
                    if pd.isna(out):
                        print('Warning: Substation {} is assigned to Zero due to no outage mentioned.'.format(start_df.iloc[index,1]))
                        start_df.iloc[index,2] = 0.0
                    elif out not in scenariolist:
                        print('ERROR: Outage Scenario {} is not covered in scenario list. Check Data Input'.format(out))
                        flag = True
                
                # decision point
                if flag:
                    return False

            else:
                wb.remove(wb[sheet_name])
        
        # Save changes
        wb.save(excel_file)
        

    except Exception as e:
        print("(Close the Excel file and Start Again) Error:", e)
        return False
    
    return start_df,oslo_total,scenariolist

# rating data frame
def check_reading_frame(sheet):
    table_start_row = 11

    print("Check Data Entry & Scenario List...")
    table_row = 12
    table_start_column = 2
    table_end_column = 11
    # create ole list and rating data
    # find total number of substations
    index = table_row
    column = table_start_column
    if sheet.cell(row=index, column=column).value is not None:
        while True:          
            index += 1
            check = sheet.cell(row=index, column=column).value
            if check is None:
                table_row_end = index
                oslo_total = index - table_row
                break
    else:
        print("Wrong data format. No information at B12")
        return False
    
    scenariolist = {}
    
    table_row = 12
    table_start_column = 22
    index = table_row
    column = table_start_column
    if sheet.cell(row=index, column=column).value is not None:
        while True:
            s_temp = sheet.cell(row=index, column=column).value
            sim_temp = sheet.cell(row=index, column=column+1).value
            file_temp = sheet.cell(row=index, column=column+2).value
            scenariolist[s_temp] = [sim_temp, file_temp]
            
            index += 1
            check = sheet.cell(row=index, column=column).value
            if check is None:
                break
        # Check if key '0' is not in scenariolist and add it to the beginning
        if 0 not in scenariolist:
            scenariolist[0] = [None, None]

    else:
        print("Wrong data format. No information at V12")
        return False
    
    return table_start_row,table_row_end,oslo_total,scenariolist

# reading essential files and save
def read_all_files(start_df,oslo_total,scenariolist,time_windows):

    result_header = ['OSLO'] + time_windows
    csv_columns = list(range(1,len(result_header)+1))

    r_df_sum = [] # result summary
    t_df_sum = [] # time summary


    # Reading result file in each folder
    for key,value in scenariolist.items():
        folder = value[0]
        sim = value[1]
        if folder == None or sim == None:
            print("WARNING: Scenario List Information Not Complete. Skipping...")
        else:            
            sim = sim + "_RMSCurrent_Sum.csv"
            file_path = os.path.join(os.getcwd(),folder,sim)

            result_df = pd.read_csv(file_path,usecols=csv_columns,nrows=oslo_total,header = None, skiprows=2, names= result_header)
            # Divide all columns by 1000 except the first column
            result_df.iloc[:, 1:] = result_df.iloc[:, 1:].divide(1000)
            r_df_sum.append(result_df)   

    
    for time in time_windows:
        print('Processing {} information in all folders...'.format(time))
        columns = ['OSLO','Substation','Outage in Scenario No.','N-0 Rating (kA)','N-1 Rating (kA)']
        for key in scenariolist.keys():
            out = "outage_scenario_" + str(key)
            columns.append(out)

        t_df = pd.DataFrame(columns=columns) # 
        
        # copy essential info
        t_df['OSLO'] = start_df['OSLO']
        t_df['Substation'] = start_df['Substation Name']
        t_df['Outage in Scenario No.'] = start_df['Outage Scenario']
        
        # copy N-0 rating
        t_df['N-0 Rating (kA)'] = start_df[time]
        time1 = time + '_out' # required due to the duplication of titile name
        t_df['N-1 Rating (kA)'] = start_df[time1]

        t_df_sum.append(t_df)


    return t_df_sum, r_df_sum

# update result table based on result
def result_update(start_df,oslo_total,scenariolist,time_windows,t_df_sum,r_df_sum):
    # r_df
    #     OSLO         1min         4min         5min        30min        60min  120min  180min
    # 0    PRLY  6838.572292  4319.651351  4216.858674  3180.263272  3047.397089     NaN     NaN
    # 1    COUL  3957.541043  3088.845394  2958.794839  2281.910857  2199.222286     NaN     NaN
    # 2    SHEP  3151.558321  2340.795600  2306.909161  1668.882423  1581.344310     NaN     NaN
    # t_df
    # OSLO    Substation  Outage in Scenario No.  N-0 Rating (kA)  N-1 Rating (kA) outage_scenario_0 outage_scenario_1 outage_scenario_2
    # 0  COUL  Substation A           1        80.000000        70.000000      NaN      NaN      NaN
    # 1  SHEP  Substation A           1        61.264000        40.842667      NaN      NaN      NaN
    # 2  HOLE  Substation A           2        40.842667        20.421333      NaN      NaN      NaN
    # 3  REDA  Substation A           2        30.632000        15.316000      NaN      NaN      NaN
    print('Updating the Final Result...')
    
    for index, df in enumerate(t_df_sum):
        time = time_windows[index]
        rindex = 0 # result index
        for key,value in scenariolist.items():
            folder = value[0]
            sim = value[1]
            if folder == None or sim == None:
                pass
                #print("WARNING: Scenario List Information Not Complete. Skipping...")
            else:            
                # merge dataframe sbased on the OSLO columns
                merged_df = pd.merge(df, r_df_sum[rindex][['OSLO', time]], left_on='OSLO', right_on='OSLO', how='left') # joing to left, left join
                rindex = rindex + 1
                # update outage_scenario column
                out = 'outage_scenario_' + str(key)
                # update the column in df based on r_df
                df[out] = merged_df[time]
                
        # update the result summary
        t_df_sum[index] = df
    
    
    return t_df_sum

# check failure and return dataframe of failures:
def check_failure(t_df_sum,scenariolist):
    print('Checking Failure Summary...')
    fail_df_sum = []
    remain = 5

    for df in t_df_sum:
        new_rows = []
        # Iterate over rows using index
        for index in range(len(df)):
            row = df.iloc[index]
            condition = False

            for scn, (key, value) in enumerate(scenariolist.items()):
                if key == row['Outage in Scenario No.']:
                    rating_column = 'N-1 Rating (kA)'
                else:
                    rating_column = 'N-0 Rating (kA)'
                
                # Use NumPy for element-wise comparison
                if row.iloc[scn + remain] >= row[rating_column]:
                    condition = True
                    break  # No need to continue checking other scenarios

            if condition:
                new_rows.append(row)

        # Create a new DataFrame at the end
        new_df = pd.DataFrame(new_rows, columns=df.columns)

        # Append the new_df to fail_df_sum
        fail_df_sum.append(new_df)


    return fail_df_sum

# group by substations:
def substation_group(fail_df_sum,scenariolist,time_windows):
    sub_df_result = []

    columns = ['Substation','N-0']
    columns = columns + time_windows
    columns = columns + ['N-1']
    columns = columns + time_windows

    rowlist = []

    for df in fail_df_sum:
        if not df.empty:  # Check if the DataFrame is not empty
            for _, row in df.iterrows():
                rowlist.append(row.tolist())

    return sub_df_result



# check the range
def get_result_range(start,space,oslo_total,scenariolist,fail_df_sum):

    range_list = []
    fix_width = 5 # 5 default columns [id, sub,out, n-0,n-1]
    columns = fix_width + len(scenariolist) 

    # 0: indivisual table frame range
    # 1: individual table data range (2 digit range)
    # 2: indivisual title range 


    # 3: result table frame range(multiple)
    # 4: result 2 digit range
    # 5: result title range

    # 6: Result  title location 
    # 7: Result  conditional formatting location [[start_row, end_row, start_column, end_column],[],[],...]

    # 8: individual  conditional formatting location [start_row, end_row, start_column, end_column]

    
    # 0
    result_range = [f"{chr(64 + 1)}{start}:{chr(64 + columns)}{start+oslo_total}"]
    range_list.append(result_range)

    # 1
    data_range = [f"{chr(64 + 1 + fix_width - 2)}{start+1}:{chr(64 + columns)}{start+oslo_total}"]
    range_list.append(data_range)
   
    # 2
    title_range = [f"{chr(64 + 1)}{start}:{chr(64 + columns)}{start}"]
    range_list.append(title_range)

    # 3 4 5
    range_list.append([])
    range_list.append([])
    range_list.append([])
    r_start = start + 2
    for index, dflist in enumerate(fail_df_sum):
        length = dflist.shape[0]
        if not length == 0:
            range_list[3].append(f"{chr(64 + 2)}{r_start+1}:{chr(64 + 1 + columns)}{r_start+1+length}")
            range_list[4].append(f"{chr(64 + 2 + fix_width - 2)}{r_start+2}:{chr(64 + 1 + columns)}{r_start+1+length}")
        
        range_list[5].append(f"{chr(64 + 2)}{r_start+1}:{chr(64 + 1 + columns)}{r_start+1}")
        r_start = r_start + length + space
        
    # result title
    range_list.append([])
    r_start = start + 2
    for index, dflist in enumerate(fail_df_sum):
        length = dflist.shape[0]
        a = f"{chr(64 + 2)}{r_start-1}" # summary title
        b = f"{chr(64 + 2 + 1)}{r_start-1}" # xxx min rms curret
        c = f"{chr(64 + 2)}{r_start}" # result
        d = f"{chr(64 + 2 + 1)}{r_start}" # pass / fail

        if not length == 0:
            e = 'Failure Substation as Table Below'
        else:
            e = 'Assessment All Pass'

        r_start = r_start + length + space
        range_list[6].append([a,b,c,d,e])

    
    # result table  location
    range_list.append([])
    r_start = start + 2
    start_row, end_row = r_start+2, r_start+1+length
    start_col, end_col = 2+fix_width, 1+columns
    
    for index, dflist in enumerate(fail_df_sum):
        length = dflist.shape[0]
        if not length == 0:
            start_row, end_row = r_start+2, r_start+1+length
            start_col, end_col = 2+fix_width, 1+columns
            range_list[7].append([start_row,end_row,start_col,end_col])
        
        r_start = r_start + length + space

        

    
    # individual section
    start_row, end_row = start+1, start+oslo_total
    start_col, end_col = 1+fix_width, columns
    range_list.append([start_row,end_row,start_col,end_col])


    return range_list

# Result table table formating
def table_formatting(simname,wb,range_list,time_windows,scenariolist):

    print("Formatting Process ...")
    # wb = load_workbook(excel_file)
    for time in time_windows:
        print("Formatting {} Result...".format(time))
        sheet = wb[time]

        sheet['A2'].value = "Result:"
        sheet['B2'].value = "Substation " + time + " RMS Current Summary"
        sheet['A3'].value = "Pass:"
        sheet['A4'].value = "Failure:"
        
        
        # Define a custom style for formatting with two decimal places
        for range_name in range_list[0]+range_list[1]:
            for row in sheet[range_name]:
                for cell in row:
                    # Apply border to all sides of the cell
                    cell.border = Border(left=Side(border_style='thin'),
                                        right=Side(border_style='thin'),
                                        top=Side(border_style='thin'),
                                        bottom=Side(border_style='thin'))
                    
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # apply_title(sheet,range_list)
        
        #print("Apply Numbering Format ...")
        # Define a custom style for formatting with two decimal places
        for range_name in range_list[1]:
            for row in sheet[range_name]:
                for cell in row:
                    cell.number_format = '0.00'


        #print("Apply Font and Shading...")
        # Shade the range B11:Z11 with a light gray background color
        for range_name in range_list[2]:
            for row in sheet[range_name]:
                for cell in row:
                    cell.font = Font(bold = True, italic = True, size = 11)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        

        condtional_formating(sheet,range_list,scenariolist,1)


        # Auto-size columns after applying formatting
        for col_letter in ['A','B',"C","D","E"]:
            sheet.column_dimensions[col_letter].auto_size = True
    
    # format the result table
    sheet = wb['Start']
    project_name = sheet['B2'].value
    feeding_desp = sheet['B4'].value
    modeller = sheet['B5'].value
    date = sheet['B6'].value

    # Result Tab process
    sheet = wb["Result"]
    sheet['B2'].value = "Project Name:"
    sheet['C2'].value = project_name
    sheet['B3'].value = "Simulation Name:"
    sheet['C3'].value = simname
    sheet['B4'].value = "Feeding Arrangement:"
    sheet['C4'].value = feeding_desp
    sheet['B5'].value = "Result Created by:"
    sheet['C5'].value = modeller
    sheet['B6'].value = "Result Created at:"
    sheet['C6'].value = datetime.now().strftime("%d-%m-%Y %H:%M")

    # sheet['J11'].value = "30min"
    # sheet['K11'].value = "time at"
    # sheet['L11'].value = "20min"
    # sheet['M11'].value = "time at"
    # sheet['N11'].value = "10min"
    # sheet['O11'].value = "time at"
    # sheet['P11'].value = "5min"
    # sheet['Q11'].value = "time at"

    # sheet['H4'].value = "Note:"
    # sheet['H5'].value = 
    # sheet['H6'].value = 

    # Define a custom style for formatting with two decimal places
    for range_name in range_list[3]+range_list[4]:
        for row in sheet[range_name]:
            for cell in row:
                # Apply border to all sides of the cell
                cell.border = Border(left=Side(border_style='thin'),
                                    right=Side(border_style='thin'),
                                    top=Side(border_style='thin'),
                                    bottom=Side(border_style='thin'))
                
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    apply_title(sheet,range_list,time_windows)
    
    print("Apply Numbering Format ...")
    # Define a custom style for formatting with two decimal places
    for range_name in range_list[4]:
        for row in sheet[range_name]:
            for cell in row:
                cell.number_format = '0.00'


    print("Apply Font and Shading...")
    # Shade the range B11:Z11 with a light gray background color
    for range_name in range_list[5]:
        for row in sheet[range_name]:
            for cell in row:
                cell.font = Font(bold = True, italic = True, size = 11)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    print("Apply Conditional Formatting ...")
    condtional_formating(sheet,range_list,scenariolist,2)
    
    #print("Apply Column Length ...")
    # Auto-adjust the width of column B based on the content in B2 to B6
    max_length = max(len(str(sheet.cell(row=i, column=2).value)) for i in range(2, 8))
    sheet.column_dimensions['B'].width = max_length + 2  # Add a little extra space


    # Auto-size columns after applying formatting
    for index in range(2,6):
        col_letter = f'{chr(64 + index)}'
        sheet.column_dimensions[col_letter].auto_size = True

    
    return

def apply_title(sheet,range_list,time_windows):

    print("Apply Title ...")

    for index, time in enumerate(time_windows):
        cell = sheet[range_list[6][index][0]]
        cell.value = "Summary Title"
        cell.font = Font(bold=True)

        cell = sheet[range_list[6][index][1]]
        cell.value = time + " TRU Assessment (RMS Current)"

        cell = sheet[range_list[6][index][2]]
        cell.value = "Result"
        cell.font = Font(bold=True)

        cell = sheet[range_list[6][index][3]]
        cell.value = range_list[6][index][4]

    return

# conditional formatting
def condtional_formating(sheet,range_list,scenariolist,option):
    # Compare values in columns I and J for each row and shade accordingly
        # set the pattern fill
    # red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red (0% S)
    # yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 255,152,51 (80% S)
    # green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green (00% S)

    key = list(scenariolist.keys())
    if option == 2:
        checking = range_list[7]
    else:
        checking = [range_list[8]]

    for index, lst in enumerate(checking):
        for r in range(lst[0],lst[1]+1):
            for c in range(lst[2],lst[3]+1):
                cell_c = sheet.cell(row = r, column = lst[2]-3)
                cell = sheet.cell(row = r, column = c)
                if cell_c.value == key[c-lst[2]]:
                    compare = sheet.cell(row = r, column = lst[2]-1)
                else:
                    compare = sheet.cell(row = r, column = lst[2]-2)
                
                if cell.value != '' and compare.value != '' and cell.value > compare.value:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    #     for scn, (key, value) in enumerate(scenariolist.items()):
    #         cell = sheet.cell(row=range_list[7][index][0],column=range_list[7][index][3])
    #         if key == row['Outage in Scenario No.']:
    #             rating_column = 'N-1 Rating (kA)'
    #         else:
    #             rating_column = 'N-0 Rating (kA)'
            
    #         # Use NumPy for element-wise comparison
    #         if row.iloc[scn + remain] >= row[rating_column]:
    #             condition = True
    #             break  # No need to continue checking other scenarios
    
    # # general
    # sheet.conditional_formatting.add(range_list[4][1],CellIsRule(operator = 'greaterThanOrEqual',formula=[range_list[4][0]],fill=red_fill))
    # sheet.conditional_formatting.add(range_list[4][1],CellIsRule(operator = 'between',formula=[range_list[4][0]+'*0.9',range_list[4][0]],fill=yellow_fill))
    # sheet.conditional_formatting.add(range_list[4][1],CellIsRule(operator = 'lessThan',formula=[range_list[4][0]],fill=green_fill))

    # for index, lst in enumerate(range_list):
    #     if index > 4:
    #         sheet.conditional_formatting.add(lst[1],CellIsRule(operator = 'greaterThanOrEqual',formula=[lst[0]],fill=red_fill))
    #         sheet.conditional_formatting.add(lst[1],CellIsRule(operator = 'between',formula=[lst[0]+'*0.9',lst[0]],fill=yellow_fill))
    #         sheet.conditional_formatting.add(lst[1],CellIsRule(operator = 'lessThan',formula=[lst[0]],fill=green_fill))


    return


if __name__ == "__main__":
    # Add your debugging code here
    simname = "DC000"  # Provide a simulation name or adjust as needed
    main_option = "1"  # Adjust as needed
    time_start = "0070000"  # Adjust as needed
    time_end = "0080000"  # Adjust as needed
    option_select = "1"  # Adjust as needed
    text_input = "test1"  # Adjust as needed
    low_v = None  # Adjust as needed
    high_v = None  # Adjust as needed
    time_step = "1"  # Adjust as needed

    # Call your main function with the provided parameters
    main(simname, main_option, time_start, time_end, option_select, text_input, low_v, high_v, time_step)

